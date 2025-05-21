import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io

def create_excel_report(company_name, email_data, linkedin_data, facebook_data, search_data, display_data):
    wb = Workbook() # Creates a workbook with one default sheet
    
    # Define styles
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    content_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    def setup_sheet_with_headers_and_styles(ws, sheet_title, columns_list, data_list):
        ws.title = sheet_title
        
        # 1. Write and style headers first
        for col_idx, column_name in enumerate(columns_list, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 2. Append data rows (will start from row 2)
        if data_list:
            df = pd.DataFrame(data_list, columns=columns_list) # Ensure consistent column order if data is dicts
            for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2): # Start r_idx at 2 for row numbers
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    # Apply content styling here directly
                    cell.alignment = content_alignment
                    cell.border = thin_border
        
        # 3. Style content cells (if data was added) and set column widths
        # This part is slightly redundant if styling is done during cell creation, but good for overall formatting
        max_rows_to_style = ws.max_row if data_list else 1 # Only style content if data exists
        for row_idx in range(2, max_rows_to_style + 1): # Iterate from row 2
            for col_idx in range(1, len(columns_list) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None: # Apply only if cell has content
                    cell.alignment = content_alignment
                    cell.border = thin_border
        
        for col_idx_letter, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            # Check header length first
            if ws.cell(row=1, column=col_idx_letter).value:
                 max_length = len(str(ws.cell(row=1, column=col_idx_letter).value))

            for cell in column_cells: # column_cells includes the header cell
                if cell.row == 1: # Skip header for max_length calculation if already considered or style differently
                    pass
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 if max_length > 0 else 15 # Ensure a minimum width
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(adjusted_width, 15), 50)


    # Email Page
    ws_email = wb.active # Get the first default sheet
    email_cols = ["Ad Name", "Funnel Stage", "Headline", "Subject Line", "Body", "CTA"]
    setup_sheet_with_headers_and_styles(ws_email, "Email", email_cols, email_data)

    # LinkedIn Page
    ws_linkedin = wb.create_sheet("LinkedIn")
    linkedin_cols = ["Ad Name", "Funnel Stage", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
    setup_sheet_with_headers_and_styles(ws_linkedin, "LinkedIn", linkedin_cols, linkedin_data)

    # FaceBook Page
    ws_facebook = wb.create_sheet("FaceBook")
    facebook_cols = ["Ad Name", "Funnel Stage", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
    setup_sheet_with_headers_and_styles(ws_facebook, "FaceBook", facebook_cols, facebook_data)

    # Google Search Page
    ws_search = wb.create_sheet("Google Search")
    search_cols = ["Headline", "Description"]
    setup_sheet_with_headers_and_styles(ws_search, "Google Search", search_cols, search_data)
    
    # Google Display Page
    ws_display = wb.create_sheet("Google Display")
    display_cols = ["Headline", "Description"]
    setup_sheet_with_headers_and_styles(ws_display, "Google Display", display_cols, display_data)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream